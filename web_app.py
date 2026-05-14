import streamlit as st
import pandas as pd
import requests
import random
from io import BytesIO
from datetime import datetime
from pathlib import Path
import sys
import openpyxl
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).parent))
from scrape_brands import (
    scrape_epicenter, scrape_eva, scrape_organic, node_available,
    write_store_sheet, write_summary_sheet, check_data_quality,
)

st.set_page_config(
    page_title="Brand Stock Checker",
    page_icon="🔍",
    layout="wide",
)

# ── Helpers ───────────────────────────────────────────────

def make_excel(results, brand, checked_at):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Summary")
    write_summary_sheet(ws, results, brand, checked_at)
    for store, products in results.items():
        ws = wb.create_sheet(store)
        if products:
            write_store_sheet(ws, products, store, brand, checked_at)
        else:
            ws["A1"] = f'No products found for "{brand}" on {store}.'
            ws["A1"].font = Font(bold=True, color="FF0000")
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def to_float(val):
    try:
        return float(str(val).replace(" ", "").replace(",", "."))
    except (ValueError, AttributeError):
        return None


def to_df(products):
    rows = []
    for p in products:
        rows.append({
            "Product Name":   p.get("name", ""),
            "SKU":            p.get("sku", ""),
            "Regular Price":  to_float(p.get("price", "")),
            "On Discount":    "Yes" if p.get("on_discount") else "No",
            "Discount Price": to_float(p.get("discount_price")) if p.get("on_discount") else None,
            "In Stock":       "Yes" if p.get("in_stock") is True else ("No" if p.get("in_stock") is False else ("Expected" if p.get("in_stock") == "expected" else "?")),
            "Seller":         p.get("seller", ""),
            "URL":            p.get("url", ""),
        })
    return pd.DataFrame(rows)


# ── Sidebar ───────────────────────────────────────────────

with st.sidebar:
    st.title("🔍 Brand Stock Checker")
    st.caption("Check product availability and prices across online stores")
    st.divider()

    st.subheader("1. Select stores")
    use_epicenter = st.checkbox("Epicenter (epicentrk.ua)",          value=True)
    use_eva       = st.checkbox("Eva (eva.ua)",                      value=True)
    use_organic   = st.checkbox("Organic Market (organic-market.com.ua)", value=True)

    st.divider()
    st.subheader("2. Select or enter brand")

    PRESET_BRANDS = [
        "Paclan", "Vileda", "Domi", "Фрекен Бок", "FINO",
        "Stella", "Spontex", "PRO SERVIS",
        "Добра Господарка", "York", "Помічниця",
        "— Custom —",
    ]

    selected_brand = st.selectbox(
        label="brand_select",
        options=PRESET_BRANDS,
        label_visibility="collapsed",
    )

    if selected_brand == "— Custom —":
        brand = st.text_input(
            label="brand_custom",
            placeholder="Type brand name...",
            label_visibility="collapsed",
        )
    else:
        brand = selected_brand

    st.divider()
    go = st.button("🔎  Search", type="primary", use_container_width=True)

    st.divider()
    if not node_available():
        st.warning(
            "⚠️ Node.js not found\n\n"
            "Epicenter scraping requires Node.js.\n"
            "Install from [nodejs.org](https://nodejs.org/)"
        )

# ── Main ──────────────────────────────────────────────────

st.title("Brand Stock Checker")

if not go:
    st.info(
        "👈 Select stores and enter a brand name in the sidebar, then click **Search**.\n\n"
        "The tool will scan all pages for that brand on each store and show a live preview "
        "with stock status, prices, and discounts. You can download the results as Excel."
    )
    st.stop()

# Validate inputs
if not brand.strip():
    st.error("Please enter a brand name.")
    st.stop()

selected = []
if use_epicenter: selected.append("Epicenter")
if use_eva:       selected.append("Eva")
if use_organic:   selected.append("Organic Market")
if not selected:
    st.error("Please select at least one store.")
    st.stop()

# ── Scraping ──────────────────────────────────────────────

has_node   = node_available()
session    = requests.Session()
results    = {}
metas      = {}
checked_at = datetime.now().strftime("%d.%m.%Y %H:%M")

for store in selected:
    with st.status(f"Scraping {store}...", expanded=True) as s:
        meta = {}
        if store == "Epicenter":
            products = scrape_epicenter(
                brand.strip(), session, has_node, log_fn=st.write, meta=meta
            )
        elif store == "Eva":
            products = scrape_eva(
                brand.strip(), session, log_fn=st.write, meta=meta
            )
        else:
            products = scrape_organic(
                brand.strip(), session, log_fn=st.write, meta=meta
            )

        results[store] = products
        metas[store]   = meta
        n      = len(products)
        in_n   = sum(1 for p in products if p["in_stock"] is True)
        disc_n = sum(1 for p in products if p.get("on_discount"))

        if n:
            s.update(
                label=f"✅ {store}: {n} products — {in_n} in stock, {disc_n} on discount",
                state="complete",
                expanded=False,
            )
        else:
            s.update(
                label=f"⚠️ {store}: no products found for '{brand.strip()}'",
                state="error",
                expanded=True,
            )

# ── Check for results ─────────────────────────────────────

total = sum(len(v) for v in results.values())
if total == 0:
    st.warning(f"No products found for **{brand.strip()}** in any selected store.")
    st.stop()

# ── Metrics ───────────────────────────────────────────────

st.divider()
total_in   = sum(sum(1 for p in v if p["in_stock"] is True)          for v in results.values())
total_out  = sum(sum(1 for p in v if p["in_stock"] is False)         for v in results.values())
total_exp  = sum(sum(1 for p in v if p["in_stock"] == "expected")    for v in results.values())
total_disc = sum(sum(1 for p in v if p.get("on_discount"))           for v in results.values())

c1, c2, c3, c4, c5 = st.columns(5)
c1.metric("Total Products", total)
c2.metric("In Stock",       total_in)
c3.metric("Out of Stock",   total_out)
c4.metric("Expected",     total_exp)
c5.metric("On Discount",    total_disc)

# ── Data Quality ──────────────────────────────────────────

st.divider()
any_issue = False
for store, products in results.items():
    meta       = metas.get(store, {})
    site_total = meta.get("site_total")
    scraped    = len(products)
    warnings   = check_data_quality(products)

    count_ok   = site_total is None or scraped >= site_total
    data_ok    = len(warnings) == 0

    if site_total or warnings:
        any_issue = any_issue or not count_ok or not data_ok
        with st.expander(
            f"{'✅' if count_ok and data_ok else '⚠️'} Data quality — {store}",
            expanded=not (count_ok and data_ok),
        ):
            if site_total:
                if count_ok:
                    st.success(f"Count check: found {scraped} products, site reports {site_total} ✅")
                else:
                    st.warning(f"Count check: found {scraped} products, but site reports {site_total} — some may be missing ⚠️")
            if warnings:
                st.warning(f"{len(warnings)} suspicious product(s) detected:")
                for msg, url in warnings:
                    if url:
                        st.caption(f"• {msg} — [Open ↗]({url})")
                    else:
                        st.caption(f"• {msg}")

            sample = random.sample(products, min(5, len(products)))
            st.markdown("**Spot-check sample** — verify these 5 products manually:")
            spot_rows = []
            for p in sample:
                spot_rows.append({
                    "Product Name":  p.get("name", ""),
                    "Regular Price": to_float(p.get("price", "")),
                    "Discount Price": to_float(p.get("discount_price")) if p.get("on_discount") else None,
                    "In Stock":      "Yes" if p.get("in_stock") is True else ("No" if p.get("in_stock") is False else ("Expected" if p.get("in_stock") == "expected" else "?")),
                    "Seller":        p.get("seller", ""),
                    "URL":           p.get("url", ""),
                })
            st.dataframe(
                pd.DataFrame(spot_rows),
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Regular Price":  st.column_config.NumberColumn("Regular Price (UAH)", format="%.2f"),
                    "Discount Price": st.column_config.NumberColumn("Discount Price (UAH)", format="%.2f"),
                    "URL": st.column_config.LinkColumn("Link", display_text="Open ↗"),
                },
            )

# ── Preview table ─────────────────────────────────────────

st.divider()

stores_with_data = [s for s in results if results[s]]
tab_labels = stores_with_data + ["📋 Summary"]
tabs = st.tabs(tab_labels)

for i, store in enumerate(stores_with_data):
    with tabs[i]:
        df = to_df(results[store])

        # Filter controls
        col_search, col_stock, col_disc = st.columns([3, 1, 1])
        with col_search:
            search_text = st.text_input(
                "Filter by name",
                key=f"search_{store}",
                placeholder="Type to filter...",
                label_visibility="collapsed",
            )
        with col_stock:
            stock_filter = st.selectbox(
                "In Stock",
                ["All", "Yes", "No", "Expected"],
                key=f"stock_{store}",
            )
        with col_disc:
            disc_filter = st.selectbox(
                "On Discount",
                ["All", "Yes", "No"],
                key=f"disc_{store}",
            )

        filtered = df.copy()
        if search_text:
            filtered = filtered[
                filtered["Product Name"].str.contains(search_text, case=False, na=False)
            ]
        if stock_filter != "All":
            filtered = filtered[filtered["In Stock"] == stock_filter]
        if disc_filter != "All":
            filtered = filtered[filtered["On Discount"] == disc_filter]

        st.caption(f"Showing {len(filtered)} of {len(df)} products")

        st.dataframe(
            filtered,
            use_container_width=True,
            hide_index=True,
            height=500,
            column_config={
                "Regular Price":  st.column_config.NumberColumn(
                    "Regular Price (UAH)", format="%.2f"
                ),
                "Discount Price": st.column_config.NumberColumn(
                    "Discount Price (UAH)", format="%.2f"
                ),
                "URL": st.column_config.LinkColumn(
                    "Product URL", display_text="Open ↗"
                ),
            },
        )

with tabs[-1]:
    rows = []
    for store, prods in results.items():
        in_n   = sum(1 for p in prods if p["in_stock"] is True)
        out_n  = sum(1 for p in prods if p["in_stock"] is False)
        exp_n  = sum(1 for p in prods if p["in_stock"] == "expected")
        disc_n = sum(1 for p in prods if p.get("on_discount"))
        rows.append({
            "Store":          store,
            "Total Products": len(prods),
            "In Stock":       in_n,
            "Out of Stock":   out_n,
            "Expected":     exp_n,
            "On Discount":    disc_n,
        })
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

# ── Download ──────────────────────────────────────────────

st.divider()
ts    = datetime.now().strftime("%Y%m%d_%H%M")
fname = f"{brand.strip().replace(' ', '_')}_{ts}.xlsx"

st.download_button(
    label="📥  Download Excel",
    data=make_excel(results, brand.strip(), checked_at),
    file_name=fname,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary",
    use_container_width=True,
)
