#!/usr/bin/env python3
"""
Brand Stock Scraper
Discovers all products of a brand on Epicenter and Eva.ua,
exports results to a multi-sheet Excel workbook.
"""

import math
import requests
import subprocess
import time
import json
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

try:
    from bs4 import BeautifulSoup
except ImportError:
    print("ERROR: Missing package. Please run START_BRANDS.bat to install dependencies.")
    input("Press Enter to close...")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: Missing package. Please run START_BRANDS.bat to install dependencies.")
    input("Press Enter to close...")
    sys.exit(1)

# ─────────────────────────────────────────────────────────
#  SETTINGS
# ─────────────────────────────────────────────────────────
DEFAULT_BRAND = "Paclan"
DELAY_SECONDS = 1.5
MAX_PAGES     = 30
# ─────────────────────────────────────────────────────────

REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "uk-UA,uk;q=0.9,ru;q=0.8,en;q=0.6",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

XLSX_COLS = [
    ("#",               5),
    ("Product Name",   55),
    ("SKU / Article",  18),
    ("Regular Price",  14),
    ("On Discount",    12),
    ("Discount Price", 14),
    ("In Stock",       11),
    ("Seller",         22),
    ("Product URL",    65),
]

C_HEADER    = "2E75B6"
C_IN_STOCK  = "E2EFDA"
C_OUT_STOCK = "FCE4D6"
C_EXPECTED  = "FFF2CC"
C_UNKNOWN   = "EDEDED"
C_TITLE     = "1F4E79"
C_DISCOUNT  = "FCE4C4"


# ─────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────

def fetch(url, session):
    try:
        r = session.get(url, headers=REQUEST_HEADERS, timeout=25, allow_redirects=True)
        r.raise_for_status()
        r.encoding = 'utf-8'
        return BeautifulSoup(r.text, "html.parser"), r.text
    except requests.RequestException as e:
        print(f"  WARN: {e}")
        return None, None


def _node_cmd():
    """Return 'node' or 'nodejs' — whichever is available."""
    for cmd in ("node", "nodejs"):
        try:
            if subprocess.run([cmd, "--version"], capture_output=True, timeout=5).returncode == 0:
                return cmd
        except (FileNotFoundError, subprocess.TimeoutExpired, OSError):
            pass
    return None


def node_available():
    return _node_cmd() is not None


def decode_nuxt(script_text, log_fn=print):
    """Run the __NUXT__ script in Node.js and return the parsed state dict."""
    import tempfile, os
    cmd = _node_cmd()
    if not cmd:
        return None
    js = "const window={};\n" + script_text + "\nprocess.stdout.write(JSON.stringify(window.__NUXT__||null));"
    tmp = None
    try:
        with tempfile.NamedTemporaryFile("w", suffix=".js", delete=False) as f:
            f.write(js)
            tmp = f.name
        result = subprocess.run(
            [cmd, tmp],
            capture_output=True, text=True, timeout=30
        )
        if result.returncode == 0 and result.stdout:
            return json.loads(result.stdout)
    except Exception:
        pass
    finally:
        if tmp:
            try: os.unlink(tmp)
            except OSError: pass
    return None


# ─────────────────────────────────────────────────────────
#  EPICENTER SCRAPER
# ─────────────────────────────────────────────────────────

def check_epicenter_product_stock(url, session):
    """Fetch a single Epicenter product page and return True/False/None."""
    soup, raw = fetch(url, session)
    if not soup:
        return None
    nuxt_data = None
    for script in soup.find_all("script"):
        txt = script.string or ""
        if "window.__NUXT__" in txt:
            nuxt_data = decode_nuxt(txt)
            break
    if not nuxt_data:
        return None
    try:
        status = nuxt_data["state"]["pages"]["shop"]["product"]["params"]["availabilityStatus"]
        code = status.get("code")
        if code == 400:
            return False
        elif code is not None:
            return True
    except (KeyError, TypeError):
        pass
    return None

def find_epicenter_brand_url(brand, session, log_fn=print):
    """Find the brand page URL on Epicenter. Returns URL or None."""
    import urllib.parse
    brand_slug = brand.lower().replace(" ", "-")
    candidates = [
        f"https://epicentrk.ua/brands/{brand_slug}.html",
        f"https://epicentrk.ua/ua/brands/{brand_slug}.html",
        f"https://epicentrk.ua/brands/{urllib.parse.quote(brand.lower())}.html",
        f"https://epicentrk.ua/ua/brands/{urllib.parse.quote(brand.lower())}.html",
    ]
    for url in candidates:
        try:
            r = session.get(url, headers=REQUEST_HEADERS, timeout=15, allow_redirects=True)
            if r.status_code == 200 and "window.__NUXT__" in r.text:
                return r.url.split("?")[0]
        except Exception:
            pass
    return None


def _epicenter_page_url(brand_base, page_num, use_brand_page):
    if page_num == 1:
        return brand_base
    return f"{brand_base}?PAGEN_1={page_num}" if use_brand_page else f"{brand_base}&page={page_num}"


def _parse_epicenter_products(raw_products):
    page_products = []
    for p in raw_products:
        name      = p.get("name_ua") or p.get("name_ru") or ""
        sku       = str(p.get("id") or "")
        url_p     = p.get("url") or ""
        price_now = p.get("price") or 0
        price_old = p.get("price_old") or 0
        on_discount    = bool(price_old and price_old > price_now)
        regular_price  = str(price_old if on_discount else price_now)
        discount_price = str(price_now) if on_discount else ""
        avail = p.get("avail")
        if avail == 100:
            in_stock = True
        elif avail == 200:
            in_stock = "expected"
        elif avail in (350, 400):
            in_stock = False
        else:
            in_stock = None
        seller = p.get("seller") or ""
        if name:
            page_products.append({
                "name": name, "sku": sku,
                "price": regular_price, "on_discount": on_discount,
                "discount_price": discount_price,
                "url": url_p, "in_stock": in_stock, "seller": seller,
            })
    return page_products


def _fetch_epicenter_page(page_num, brand_base, use_brand_page, session):
    url = _epicenter_page_url(brand_base, page_num, use_brand_page)
    soup, raw = fetch(url, session)
    if not soup:
        return page_num, None
    for script in soup.find_all("script"):
        txt = script.string or ""
        if "window.__NUXT__" in txt:
            nuxt_data = decode_nuxt(txt)
            if nuxt_data:
                try:
                    raw_prods = nuxt_data["state"]["products"]["products"]
                    return page_num, _parse_epicenter_products(raw_prods)
                except (KeyError, TypeError):
                    pass
    return page_num, None


def scrape_epicenter(brand, session, has_node, log_fn=print, meta=None):
    if not has_node:
        log_fn("Epicenter: SKIPPED — Node.js is not installed. Install from https://nodejs.org/")
        return []

    log_fn(f"Epicenter: searching for '{brand}'...")

    brand_base = find_epicenter_brand_url(brand, session, log_fn)
    if brand_base:
        log_fn(f"  Brand page found: {brand_base}")
    else:
        brand_enc = requests.utils.quote(brand)
        brand_base = f"https://epicentrk.ua/ua/search/?q={brand_enc}&per-page=60"
        log_fn(f"  ⚠️ Brand page not found — using search (stock status may be incomplete)")

    use_brand_page = "brands" in brand_base

    # ── Fetch page 1 to get total_pages ──
    soup, raw = fetch(_epicenter_page_url(brand_base, 1, use_brand_page), session)
    if not soup:
        log_fn("  Page 1: failed to load")
        return []

    nuxt_data = None
    for script in soup.find_all("script"):
        txt = script.string or ""
        if "window.__NUXT__" in txt:
            nuxt_data = decode_nuxt(txt, log_fn=log_fn)
            break

    if not nuxt_data:
        log_fn("  Page 1: could not decode page data")
        return []

    try:
        pagination  = nuxt_data["data"][0]["params"]["pagination"]
        total_pages = pagination.get("pages", 1)
        site_total  = pagination.get("count") or pagination.get("total")
        if site_total and meta is not None:
            meta["site_total"] = int(site_total)
    except (KeyError, IndexError, TypeError):
        total_pages = 1

    try:
        page1_products = _parse_epicenter_products(nuxt_data["state"]["products"]["products"])
    except (KeyError, TypeError):
        page1_products = []

    log_fn(f"  Page 1/{total_pages}: {len(page1_products)} products")
    all_products = list(page1_products)

    # ── Fetch remaining pages in parallel (3 at a time) ──
    if total_pages > 1:
        remaining = list(range(2, min(total_pages, MAX_PAGES) + 1))
        with ThreadPoolExecutor(max_workers=3) as executor:
            futures = {
                executor.submit(_fetch_epicenter_page, p, brand_base, use_brand_page, session): p
                for p in remaining
            }
            page_results = {}
            for future in as_completed(futures):
                page_num, products = future.result()
                page_results[page_num] = products or []
                log_fn(f"  Page {page_num}/{total_pages}: {len(page_results[page_num])} products")

        for p in remaining:
            all_products.extend(page_results.get(p, []))

    if meta is not None:
        meta["scraped_total"] = len(all_products)
        site_total = meta.get("site_total")
        if site_total:
            if len(all_products) >= site_total:
                log_fn(f"  Count check ✅ — found {len(all_products)}, site reports {site_total}")
            else:
                log_fn(f"  Count check ⚠️ — found {len(all_products)}, site reports {site_total} (may be incomplete)")

    return all_products


# ─────────────────────────────────────────────────────────
#  EVA SCRAPER
# ─────────────────────────────────────────────────────────

def find_eva_brand_id(brand_name, session):
    """Look up Eva brand ID by name. Returns (id, title) or (None, None)."""
    try:
        r = session.get(
            "https://api.eva.ua/v1/ua/api/brands",
            headers={**REQUEST_HEADERS, "Accept": "application/json"},
            timeout=15,
        )
        data = r.json()
        for group_brands in data.get("data", {}).get("groups", {}).values():
            for b in group_brands:
                if b.get("title", "").lower() == brand_name.lower():
                    m = re.search(r"brnd-(\d+)", b.get("url", ""))
                    if m:
                        return m.group(1), b["title"]
    except Exception as e:
        print(f"  WARN: could not fetch Eva brand list: {e}")
    return None, None


def _resolve_nuxt(arr, idx, depth=0):
    """Dereference Nuxt 3 array-ref payload format."""
    if depth > 30 or not isinstance(idx, int) or idx < 0 or idx >= len(arr):
        return None
    val = arr[idx]
    if isinstance(val, list) and len(val) == 2 and isinstance(val[0], str) and isinstance(val[1], int):
        return _resolve_nuxt(arr, val[1], depth + 1)
    if isinstance(val, dict):
        return {k: (_resolve_nuxt(arr, v, depth + 1) if isinstance(v, int) else v) for k, v in val.items()}
    if isinstance(val, list):
        return [_resolve_nuxt(arr, v, depth + 1) if isinstance(v, int) else v for v in val]
    return val


def parse_eva_nuxt_payload(html_text, brand_id):
    """Extract brand data from Nuxt 3 <script type='application/json'> payload."""
    soup = BeautifulSoup(html_text, "html.parser")
    tag = soup.find("script", {"type": "application/json"})
    if not tag:
        return None
    try:
        arr = json.loads(tag.string or "")
    except (json.JSONDecodeError, TypeError):
        return None
    brand_key = f"brnd-brnd-{brand_id}"
    for el in arr:
        if isinstance(el, dict) and brand_key in el:
            idx = el[brand_key]
            if isinstance(idx, int) and idx > 0:
                return _resolve_nuxt(arr, idx)
    return None


def _parse_eva_products(brand_data):
    page_products = []
    for p in brand_data.get("hits", []):
        name   = p.get("name") or ""
        sku    = str(p.get("sku") or "")
        price  = p.get("price") or 0
        final  = p.get("final_price") or p.get("special_price") or price
        on_discount    = bool(final and price and final < price)
        regular_price  = str(price)
        discount_price = str(final) if on_discount else ""
        stock       = p.get("stock") or {}
        in_stock    = stock.get("is_in_stock")
        url_p       = f"https://eva.ua/ua/search/?q={sku}" if sku else ""
        merchant_id = p.get("merchant_id")
        seller      = "EVA" if merchant_id == 1 else (f"#{merchant_id}" if merchant_id else "")
        if name:
            page_products.append({
                "name": name, "sku": sku,
                "price": regular_price, "on_discount": on_discount,
                "discount_price": discount_price,
                "url": url_p, "in_stock": in_stock, "seller": seller,
            })
    return page_products


def _fetch_eva_page(page_num, base_url, brand_id, session):
    url = f"{base_url}?p={page_num}" if page_num > 1 else base_url
    _, raw = fetch(url, session)
    if not raw:
        return page_num, None
    brand_data = parse_eva_nuxt_payload(raw, brand_id)
    if not brand_data:
        return page_num, None
    return page_num, _parse_eva_products(brand_data)


def scrape_eva(brand, session, log_fn=print, meta=None):
    log_fn(f"Eva: searching for '{brand}'...")
    log_fn("  Looking up brand in Eva catalogue...")
    brand_id, found_title = find_eva_brand_id(brand, session)
    if not brand_id:
        log_fn(f"  Brand '{brand}' was not found in Eva's brand list.")
        return []
    log_fn(f"  Found: '{found_title}' (ID {brand_id})")

    base_url = f"https://eva.ua/ua/brnd-{brand_id}/"

    # ── Fetch page 1 to get total_pages ──
    _, raw = fetch(base_url, session)
    if not raw:
        log_fn("  Page 1: failed to load")
        return []

    brand_data = parse_eva_nuxt_payload(raw, brand_id)
    if not brand_data:
        log_fn("  Page 1: could not parse page data")
        return []

    site_total = brand_data.get("total")
    if site_total and meta is not None:
        meta["site_total"] = int(site_total)

    page1_products = _parse_eva_products(brand_data)
    per_page    = len(page1_products) if page1_products else 40
    total_pages = math.ceil(int(site_total) / per_page) if site_total and per_page else 1

    log_fn(f"  Page 1/{total_pages}: {len(page1_products)} products")
    all_products = list(page1_products)

    # ── Fetch remaining pages in parallel (3 at a time) ──
    if total_pages > 1:
        remaining = list(range(2, min(total_pages, MAX_PAGES) + 1))
        with ThreadPoolExecutor(max_workers=3) as executor:
            futures = {
                executor.submit(_fetch_eva_page, p, base_url, brand_id, session): p
                for p in remaining
            }
            page_results = {}
            for future in as_completed(futures):
                page_num, products = future.result()
                page_results[page_num] = products or []
                log_fn(f"  Page {page_num}/{total_pages}: {len(page_results[page_num])} products")

        for p in remaining:
            all_products.extend(page_results.get(p, []))

    if meta is not None:
        meta["scraped_total"] = len(all_products)
        site_total = meta.get("site_total")
        if site_total:
            if len(all_products) >= site_total:
                log_fn(f"  Count check ✅ — found {len(all_products)}, site reports {site_total}")
            else:
                log_fn(f"  Count check ⚠️ — found {len(all_products)}, site reports {site_total} (may be incomplete)")

    return all_products


# ─────────────────────────────────────────────────────────
#  DATA QUALITY CHECK
# ─────────────────────────────────────────────────────────

def check_data_quality(products):
    """Return list of (message, url) tuples for suspicious product data."""
    warnings = []
    for p in products:
        name = p.get("name", "")[:50]
        url  = p.get("url", "")
        try:
            price = float(str(p.get("price", 0)).replace(" ", "").replace(",", "."))
        except (ValueError, TypeError):
            price = 0
        try:
            disc  = float(str(p.get("discount_price", 0) or 0).replace(" ", "").replace(",", "."))
        except (ValueError, TypeError):
            disc = 0

        if price == 0 and p.get("in_stock") not in (False, "expected"):
            warnings.append((f"Price is 0: {name}", url))
        if p.get("on_discount") and disc >= price:
            warnings.append((f"Discount price ≥ regular price: {name}", url))
        if p.get("on_discount") and price > 0 and disc > 0 and (price - disc) / price > 0.9:
            warnings.append((f"Discount >90% (suspicious): {name}", url))
    return warnings


# ─────────────────────────────────────────────────────────
#  EXCEL WRITER
# ─────────────────────────────────────────────────────────

def _fmt_price(raw):
    digits = re.sub(r"[^\d.]", "", str(raw).replace(",", "."))
    try:
        return "{:,.2f}".format(float(digits)).replace(",", " ")
    except ValueError:
        return str(raw)


def write_store_sheet(ws, products, store_name, brand, checked_at):
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor=C_HEADER)
    in_fill   = PatternFill("solid", fgColor=C_IN_STOCK)
    out_fill  = PatternFill("solid", fgColor=C_OUT_STOCK)
    exp_fill  = PatternFill("solid", fgColor=C_EXPECTED)
    unk_fill  = PatternFill("solid", fgColor=C_UNKNOWN)
    center   = Alignment(horizontal="center", vertical="center")
    left     = Alignment(horizontal="left",   vertical="center")
    wrap     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    last_col = get_column_letter(len(XLSX_COLS))

    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value     = f"{store_name}  ·  Brand: {brand}  ·  Checked: {checked_at}"
    t.font      = Font(bold=True, size=13, color=C_TITLE)
    t.alignment = center
    ws.row_dimensions[1].height = 26

    for ci, (col_name, col_width) in enumerate(XLSX_COLS, 1):
        c = ws.cell(row=2, column=ci, value=col_name)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = center
        ws.column_dimensions[get_column_letter(ci)].width = col_width
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    disc_fill = PatternFill("solid", fgColor=C_DISCOUNT)
    aligns = [center, wrap, center, center, center, center, center, center, left]
    for i, p in enumerate(products, 1):
        row = i + 2
        if p["in_stock"] is True:
            status, stock_fill = "Yes", in_fill
        elif p["in_stock"] is False:
            status, stock_fill = "No", out_fill
        elif p["in_stock"] == "expected":
            status, stock_fill = "Expected", exp_fill
        else:
            status, stock_fill = "?", unk_fill

        on_disc = p.get("on_discount", False)
        vals = [
            i,
            p.get("name", ""),
            p.get("sku", ""),
            _fmt_price(p.get("price", "")),
            "Yes" if on_disc else "No",
            _fmt_price(p.get("discount_price", "")) if on_disc else "",
            status,
            p.get("seller", ""),
            p.get("url", ""),
        ]
        for ci, (val, aln) in enumerate(zip(vals, aligns), 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.alignment = aln
            if ci in (5, 6) and on_disc:
                c.fill = disc_fill
            elif ci == 7:
                c.fill = stock_fill

    fr     = len(products) + 4
    in_n   = sum(1 for p in products if p["in_stock"] is True)
    out_n  = sum(1 for p in products if p["in_stock"] is False)
    exp_n  = sum(1 for p in products if p["in_stock"] == "expected")
    unk_n  = sum(1 for p in products if p["in_stock"] is None)
    disc_n = sum(1 for p in products if p.get("on_discount"))
    bf     = Font(bold=True)
    ws.cell(row=fr, column=1, value="TOTAL").font = bf
    ws.cell(row=fr, column=2, value=f"{len(products)} products").font = bf
    ws.cell(row=fr, column=5, value=f"On discount: {disc_n}").font = bf
    ws.cell(row=fr, column=7,
            value=f"In stock: {in_n}  |  Out of stock: {out_n}  |  Expected: {exp_n}  |  Unknown: {unk_n}").font = bf


def write_summary_sheet(ws, results, brand, checked_at):
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor=C_HEADER)
    in_fill  = PatternFill("solid", fgColor=C_IN_STOCK)
    out_fill = PatternFill("solid", fgColor=C_OUT_STOCK)
    center   = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value     = f"Summary  ·  Brand: {brand}  ·  {checked_at}"
    t.font      = Font(bold=True, size=13, color=C_TITLE)
    t.alignment = center
    ws.row_dimensions[1].height = 26

    disc_fill = PatternFill("solid", fgColor=C_DISCOUNT)
    exp_fill  = PatternFill("solid", fgColor=C_EXPECTED)
    headers = ["Store", "Total Products", "In Stock", "Out of Stock", "Expected", "On Discount", "Unknown"]
    widths  = [20, 16, 12, 14, 14, 13, 12]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, (store, products) in enumerate(results.items(), 3):
        in_n   = sum(1 for p in products if p["in_stock"] is True)
        out_n  = sum(1 for p in products if p["in_stock"] is False)
        exp_n  = sum(1 for p in products if p["in_stock"] == "expected")
        unk_n  = sum(1 for p in products if p["in_stock"] is None)
        disc_n = sum(1 for p in products if p.get("on_discount"))
        for ci, val in enumerate([store, len(products), in_n, out_n, exp_n, disc_n, unk_n], 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = center
            if ci == 3 and in_n > 0:
                c.fill = in_fill
            if ci == 4 and out_n > 0:
                c.fill = out_fill
            if ci == 5 and exp_n > 0:
                c.fill = exp_fill
            if ci == 6 and disc_n > 0:
                c.fill = disc_fill


# ─────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────

def main():
    print("=" * 56)
    print("  BRAND STOCK SCRAPER")
    print("=" * 56)

    has_node = node_available()
    if not has_node:
        print("\n  WARNING: Node.js not found.")
        print("  Epicenter scraping requires Node.js.")
        print("  Install from https://nodejs.org/ then re-run.\n")

    entered = input(f"\nEnter brand name (press Enter for '{DEFAULT_BRAND}'): ").strip()
    brand = entered if entered else DEFAULT_BRAND

    checked_at  = datetime.now().strftime("%d.%m.%Y %H:%M")
    ts          = datetime.now().strftime("%Y%m%d_%H%M")
    output_file = f"{brand.replace(' ', '_')}_{ts}.xlsx"
    output_path = Path(__file__).parent / output_file

    print(f"\n  Brand:  {brand}")
    print(f"  Output: {output_file}")
    print("-" * 56)

    session = requests.Session()
    results = {}

    results["Epicenter"] = scrape_epicenter(brand, session, has_node, log_fn=print)
    results["Eva"]       = scrape_eva(brand, session, log_fn=print)

    for store, products in results.items():
        in_n = sum(1 for p in products if p["in_stock"] is True)
        print(f"  {store}: {len(products)} products, {in_n} in stock")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_sum = wb.create_sheet("Summary")
    write_summary_sheet(ws_sum, results, brand, checked_at)

    for store_name, products in results.items():
        ws = wb.create_sheet(store_name)
        if products:
            write_store_sheet(ws, products, store_name, brand, checked_at)
        else:
            ws["A1"] = f'No products found for "{brand}" on {store_name}.'
            ws["A1"].font = Font(bold=True, color="FF0000")
            if store_name == "Epicenter" and not has_node:
                ws["A2"] = "Node.js is required — install from https://nodejs.org/"
            else:
                ws["A2"] = "Check that the brand name matches exactly how it appears on the store."

    wb.save(output_path)

    total    = sum(len(v) for v in results.values())
    total_in = sum(sum(1 for p in v if p["in_stock"] is True) for v in results.values())

    print(f"\n{'=' * 56}")
    print(f"  DONE  —  {total} products total, {total_in} in stock")
    print(f"  File: {output_file}")
    print(f"{'=' * 56}")
    input("\nPress Enter to close...")


if __name__ == "__main__":
    main()
